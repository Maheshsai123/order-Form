from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from openpyxl import Workbook, load_workbook
import os, datetime, json

app = Flask(__name__)
app.secret_key = "secret"

EXCEL_FILE = "orders.xlsx"
UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

VENDORS = ["Mahesh", "Sai", "Ravi", "Other"]
doc_counter = 0

def get_next_doc_no():
    """Generate next document number automatically"""
    global doc_counter
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        max_doc = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            doc = row[0]
            if doc and doc.startswith("DOC-"):
                num = int(doc.split("-")[1])
                if num > max_doc:
                    max_doc = num
        doc_counter = max_doc
    doc_counter += 1
    return f"DOC-{doc_counter:04d}"

@app.route("/", methods=["GET", "POST"])
def client():
    """Client page for placing orders"""
    if request.method == "POST":
        doc_no = request.form.get("doc_no")
        vendor = request.form.get("vendor")
        date = request.form.get("date")
        order_json = request.form.get("order_json")

        try:
            orders = json.loads(order_json)
        except Exception:
            flash("Invalid order data.")
            return redirect(url_for("client"))

        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(["Doc No", "Vendor", "Date", "S.No", "Item", "Design", "Size", "Qty", "Image Path"])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        for i, row in enumerate(orders):
            sno, item, design, size, qty, _ = row
            image_path = ""
            file_key = f"image_{i}"
            if file_key in request.files:
                file = request.files[file_key]
                filename = f"{doc_no}_{sno}_{file.filename}"
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                file.save(filepath)
                image_path = filepath
            ws.append([doc_no, vendor, date, sno, item, design, size, qty, image_path])

        wb.save(EXCEL_FILE)
        flash(f"✅ Order {doc_no} saved successfully!")
        return redirect(url_for("client"))

    doc_no = get_next_doc_no()
    today = datetime.date.today().strftime("%Y-%m-%d")
    return render_template("base.html", doc_no=doc_no, today=today, vendors=VENDORS)

@app.route("/admin")
def admin():
    """Admin page to view orders"""
    if not os.path.exists(EXCEL_FILE):
        return render_template("admin.html", orders=[])

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    orders = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template("admin.html", orders=orders)

@app.route("/download")
def download_excel():
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True)
    else:
        return "No file found", 404

if __name__ == "__main__":
    app.run(debug=True,port=5001)

